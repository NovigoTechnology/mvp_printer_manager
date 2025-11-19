from io import BytesIO, StringIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
import csv
from typing import List, Dict, Any

class ExportService:
    def __init__(self):
        # Estilos para Excel
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2980b9", end_color="2980b9", fill_type="solid")
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_invoices_to_excel(self, invoices_data: List[Dict], period_name: str = None):
        """Exportar facturas a Excel"""
        buffer = BytesIO()
        
        # Crear DataFrame
        if not invoices_data:
            # Si no hay datos, crear DataFrame vacío con columnas predefinidas
            columns = ['Número Factura', 'ID Contrato', 'Nombre Contrato', 'Fecha Factura',
                      'Fecha Vencimiento', 'Total', 'Estado', 'Moneda']
            df = pd.DataFrame(columns=columns)
        else:
            df = pd.DataFrame(invoices_data)
            
            # Renombrar columnas para que sean más descriptivas
            column_mapping = {
                'invoice_number': 'Número Factura',
                'contract_id': 'ID Contrato',
                'contract_number': 'Número Contrato',
                'contract_name': 'Nombre Contrato',
                'supplier': 'Proveedor',
                'invoice_date': 'Fecha Factura',
                'due_date': 'Fecha Vencimiento',
                'period_start': 'Inicio Período',
                'period_end': 'Fin Período',
                'subtotal': 'Subtotal',
                'tax_amount': 'Impuestos',
                'total_amount': 'Total',
                'status': 'Estado',
                'currency': 'Moneda',
                'tax_rate': 'Tasa Impuesto',
                'notes': 'Notas'
            }
            
            df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns}, inplace=True)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Facturas{f' - {period_name}' if period_name else ''}"
        
        # Agregar título
        title = f"Reporte de Facturas{f' - {period_name}' if period_name else ''}"
        if len(df.columns) > 0:
            ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
        else:
            ws.merge_cells('A1:H1')
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = self.center_alignment
        ws['A1'].fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        
        # Agregar fecha de generación
        if len(df.columns) > 0:
            ws.merge_cells(f'A2:{get_column_letter(len(df.columns))}2')
        else:
            ws.merge_cells('A2:H2')
        ws['A2'] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = self.center_alignment
        
        # Agregar datos
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Aplicar formato solo si hay columnas
        if len(df.columns) > 0:
            self._apply_excel_formatting(ws, df, 4)
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def export_readings_to_excel(self, readings_data: List[Dict], period_name: str = None):
        """Exportar lecturas de contadores a Excel"""
        buffer = BytesIO()
        
        # Crear DataFrame
        if not readings_data:
            # Si no hay datos, crear DataFrame vacío con columnas predefinidas
            columns = ['ID Impresora', 'IP Impresora', 'Ubicación', 'ID Contrato', 'Nombre Contrato',
                      'Fecha Lectura', 'Contador B/N Actual', 'Contador Color Actual', 
                      'Contador Total Actual', 'Método Lectura', 'Notas']
            df = pd.DataFrame(columns=columns)
        else:
            df = pd.DataFrame(readings_data)
            
            # Renombrar columnas
            column_mapping = {
                'printer_id': 'ID Impresora',
                'printer_ip': 'IP Impresora',
                'printer_location': 'Ubicación',
                'contract_id': 'ID Contrato',
                'contract_name': 'Nombre Contrato',
                'reading_date': 'Fecha Lectura',
                'counter_bw_current': 'Contador B/N Actual',
                'counter_color_current': 'Contador Color Actual',
                'counter_total_current': 'Contador Total Actual',
                'counter_bw_previous': 'Contador B/N Anterior',
                'counter_color_previous': 'Contador Color Anterior',
                'counter_total_previous': 'Contador Total Anterior',
                'prints_bw_period': 'Impresiones B/N Período',
                'prints_color_period': 'Impresiones Color Período',
                'prints_total_period': 'Impresiones Total Período',
                'reading_method': 'Método Lectura',
                'notes': 'Notas'
            }
            
            df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns}, inplace=True)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Lecturas{f' - {period_name}' if period_name else ''}"
        
        # Agregar título
        title = f"Reporte de Lecturas de Contadores{f' - {period_name}' if period_name else ''}"
        if len(df.columns) > 0:
            ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
        else:
            ws.merge_cells('A1:K1')
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = self.center_alignment
        ws['A1'].fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        
        # Agregar fecha de generación
        if len(df.columns) > 0:
            ws.merge_cells(f'A2:{get_column_letter(len(df.columns))}2')
        else:
            ws.merge_cells('A2:K2')
        ws['A2'] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = self.center_alignment
        
        # Agregar datos
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Aplicar formato solo si hay columnas
        if len(df.columns) > 0:
            self._apply_excel_formatting(ws, df, 4)
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def export_financial_summary_to_excel(self, summary_data: Dict[str, Any]):
        """Exportar resumen financiero a Excel"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Financiero"
        
        # Título
        ws.merge_cells('A1:D1')
        ws['A1'] = "Resumen Financiero de Facturación"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = self.center_alignment
        ws['A1'].fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
        
        # Fecha
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = self.center_alignment
        
        # Datos del resumen
        row = 4
        metrics = [
            ('Total de Facturas', summary_data.get('total_invoices', 0)),
            ('Ingresos Totales', f"${summary_data.get('total_revenue', 0):,.2f}"),
            ('Facturas Pagadas', summary_data.get('paid_invoices', 0)),
            ('Facturas Pendientes', summary_data.get('pending_invoices', 0)),
            ('Facturas Vencidas', summary_data.get('overdue_invoices', 0)),
            ('Períodos Activos', summary_data.get('active_periods', 0)),
            ('Contratos Activos', summary_data.get('active_contracts', 0)),
        ]
        
        for metric_name, metric_value in metrics:
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = metric_value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def export_invoices_to_csv(self, invoices_data: List[Dict]):
        """Exportar facturas a CSV"""
        buffer = StringIO()
        
        if invoices_data:
            df = pd.DataFrame(invoices_data)
            df.to_csv(buffer, index=False, encoding='utf-8')
        else:
            # CSV vacío con encabezados
            buffer.write("invoice_number,contract_id,invoice_date,total_amount,status\n")
        
        output = BytesIO(buffer.getvalue().encode('utf-8'))
        return output
    
    def _apply_excel_formatting(self, ws, df, header_row):
        """Aplicar formato a una hoja de Excel"""
        # Formatear encabezados
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Formatear datos
        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajustar ancho de columnas
        for col_num in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width